from openpyxl.styles import NamedStyle, Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.colors import Color
sheet_style = {
    "col_A_width": 8.00,
    "col_B_width": 45.00,
    "col_C_width": 16.00,
    "col_D_width": 16.00,
    "col_E_width": 8.00,
    "col_F_width": 20.00,
}

info_style = NamedStyle(name="info_style")
info_style.font = Font(color="FFFFFF", bold=True, size=10)
info_style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
info_style.fill = PatternFill(fill_type="solid", start_color="00000000", end_color="00000000")

field_style = NamedStyle(name="field_style")
field_style.font = Font(color="FFFFFF", bold=True, size=10)
field_style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
field_style.fill = PatternFill(fill_type="solid", start_color="00555555", end_color="00555555")

record_style = NamedStyle(name="record_style")
thin = Side(border_style="thin", color="000000")
record_style.border = Border(top=thin, left=thin, right=thin, bottom=thin)
record_style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
record_style.font = Font(color="000000", bold=False, size=10)