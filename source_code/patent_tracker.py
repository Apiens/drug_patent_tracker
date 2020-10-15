import asyncio
import time
import os
from datetime import datetime, date, timedelta
import aiohttp
from bs4 import BeautifulSoup
from typing import Union, List, Tuple, Iterator, Iterable, Dict
import openpyxl
from default_style import info_style, field_style, record_style, sheet_style
import string

api_key_path = "../input_setting/api_key.txt"
input_path = "../input_setting/input.xlsx"
output_directory = "../output"

## FYI ##
#current version: v1.0
#Comments & Docstring: English (for developers)
#print message: Korean (for consumers)

#TODO before 2.0:
# 1. Crwaling by Selenium -> (despite it's heaviness) Works without api_key.
# 2. More output formats: to_img(), to_csv()

class PatentTracker():
    """Creates PatentTracker instance.
    Consists of 4 kinds of methods
    I. GETTERS AND SETTERS: api_key, since, before, targets, results
    II. SETTINGS AND INPUT: read_and_check_api_key, read_input
    III. TRACKING: track_patents
    IV. OUTPUT: to_excel and private methods

    -- Quick Example --
    tracker = PatentTracker()
    
    time1 = time.time()
    loop = asyncio.get_event_loop()
    loop.run_until_complete(tracker.track_patents())
    loop.close
    time2 = time.time()
    print(f"total time taken: {time2-time1}")
    
    tracker.to_excel()
    """

    def __init__(self):
        self.__api_key: str = ""
        self.__targets: Iterator = (i for i in [])
        self.__since:date = date(1970,1,1)
        self.__before:date = datetime.now().date()
        self.__additional_info_dict: Dict = {}
        self.__results: List = []

        if os.path.exists(input_path): self.read_input(input_path)
        else: raise FileNotFoundError(f"{input_path} 를 찾을 수 없습니다.")

        if os.path.exists(api_key_path): self.read_and_check_api_key(api_key_path)
        else: raise FileNotFoundError(f"{api_key_path}를 찾을 수 없습니다.")

    ############### I. GETTERS AND SETTERS ###############
    @property
    def api_key(self):
        return self.__api_key
    @api_key.setter
    def api_key(self, api_key:str):
        self.__api_key = api_key.strip()
        print(f"tracker.api_key is now {self.__api_key}")
    
    @property
    def since(self):
        return self.__since
    @since.setter
    def since(self, new_since):
        if isinstance(new_since, datetime):
            new_since = datetime.strftime(new_since, "%Y/%m/%d")
        new_since = "/".join(new_since.split(" "))
        new_since = "/".join(new_since.split("."))
        new_since = "/".join(new_since.split("-"))
        self.__since = datetime.strptime(new_since, "%Y/%m/%d").date()
        print(f"tracker.since is set as {self.__since}")
    
    @property
    def before(self):
        return self.__before
    @before.setter
    def before(self, new_before):
        if isinstance(new_before, datetime):
            new_before = datetime.strftime(new_before, "%Y/%m/%d")
        new_before = "/".join(new_before.split(" "))
        new_before = "/".join(new_before.split("."))
        new_before = "/".join(new_before.split("-"))
        self.__before = datetime.strptime(new_before, "%Y/%m/%d").date()
        print(f"tracker.before is set as {self.__before}")

    @property
    def targets(self):
        return self.__targets
    @targets.setter
    def targets(self, targets: List[str]):
        weird_input = [target for target in targets if len(target) != 13]
        if weird_input:
            raise ValueError(
                # "Some input elements does not satisfy condition." + "\n" +
                # "Please check following elements and make sure they are 13 digit" +"\n" + 
                # f"{weird_input}"
                "일부 타겟 정보가 올바르지 않습니다.\n"+
                "아래 출원번호값들을 확인 해 주세요.\n"+
                f"{weird_input}\n"+
                "출원번호는 13자리 숫자 값이어야 합니다."
                )
        targets_gen = (target if len(target)==13 else "".join(target.split("-")) for target in targets)
        self.__targets = targets_gen
        print("타겟 입력 성공.")
        # print("successfully fetched targets. ")

    @property
    def additional_info_dict(self):
        return self.__additional_info_dict

    @property
    def results(self):
        return self.__results

    ############### II. LOAD SETTINGS, INPUT ###############
    def read_and_check_api_key(self, path:str, verbose=False):
        with open(path, "r") as text:
            api_key = text.readline()
        if verbose: print("API key를 api_key.txt 로부터가져왔습니다.")
        #print(f"Read api_key from api_key.txt as : {api_key}")
        if self.check_api_key(api_key):
            self.api_key = api_key
        else:
            print("읽어온 API key가 유효하지 않습니다. 기존 값으로 유지됩니다.")
            print(f"읽어온 key 값: {api_key}")
            print(f"현재 API key: {self.__api_key}")


    #TODO: 
    def check_api_key(self, api_key):
        # print(f"checking api_key vailidy")
        # import urllib
        # url = 'http://plus.kipris.or.kr/openapi/rest/RelatedDocsonfilePatService/relatedDocsonfileInfo'
        # query = f'?applicationNumber=1019940701319&accessKey={api_key}'
        # urllib.request(url+query)
        return True

    def read_input(self, input_path:str=input_path, verbose=False):
        """Reads file that contains input info and assigns properties using setters.
        self.__targets, self.since, self.before will be assigned.

        Argument(s)
        - input_path:str = input_path
        """
        targets:List[str]  = []
        additional_info_dict:Dict = {}
        
        if os.path.exists(input_path): 
            wb = openpyxl.load_workbook(input_path) #Read xl file
            if verbose: print(f"{input_path}로 부터 인풋 값을 불러옵니다.")

            # Fetching application numbers from sheet 'targets'
            try:target_sheet = wb['targets']
            except KeyError:
                print(f"{input_path} 에 'targets' 시트가 존재하지 않습니다.")
                print("'Sheet2' 시트로 부터 타겟 정보를 불러옵니다.")
                target_sheet = wb['Sheet1']
            for row in target_sheet.iter_rows(min_row=3, max_row=5000, max_col=6, values_only=True):
                if row[0] == None: continue
                targets.append(str(row[0])) # application number.
                additional_info_dict[str(row[0])] = row[1:] # additional info of target patent.
                # print(f"출원번호 {row[0]} 및 추가정보를 읽어왔습니다.")

            self.targets = targets #saved as generator
            if verbose:
                print(f"targets[:3]: {targets[:3]}")
                print(f"self.targets: {self.targets}")
            self.__additional_info_dict = additional_info_dict
            if verbose:
                print("타겟 정보를 성공적으로 불러왔습니다.")
                print(f"타겟 수: {len(targets)}")
                print(f"첫번째 타겟 출원번호: {list(self.additional_info_dict.keys())[0]}")
                print(f"첫번째 타겟 부가정보: {list(self.additional_info_dict.values())[0]}")

            # Reading date info from sheet 'dates'
            try:dates_sheet = wb['dates']
            except KeyError:
                print(f"{input_path} 에 'dates' 시트가 존재하지 않습니다.")
                # print("Current excel file doesn't have a sheet named 'dates'")
                print(f"'Sheet2' 시트로 부터 날짜구간 정보를 불러옵니다.")
                # print("worksheet 'Sheet2' will be open instead of worksheet 'dates'")
                dates_sheet = wb['Sheet2']
            
            last_n_day = abs(dates_sheet['C2'].value)
            if last_n_day:
                self.__before = datetime.now().date()
                self.__since = self.__before - timedelta(days=last_n_day)
            else:
                self.since = dates_sheet['C3'].value
                self.before = dates_sheet['C4'].value
        else:
            print(f"{input_path} 파일이 존재하지 않습니다.")
            # print(f"file does not exist in the path: {input_path}")  

    ############### III. TRACKING ###############
    async def track_patents(self, verbose=False):
        """Asynchronously tracks patents in self.targets

        Simply operates by repeating self.track_patent()
        
        Saves a list containing tuples at self.__results.
            [(application_number, result_2D_table), (...), ...]
        """
        
        # returned values of each task will be appended into an empty list and then returned.
        futures = [asyncio.ensure_future(self.track_patent(patent, verbose=verbose)) for patent in self.targets]
        results = await asyncio.gather(*futures)

        ## this code will work synchronously -> compare with async
        # results = []
        # for patent in self.targets:
        #     results.append(await self.track_patent(patent))
        # print(results)
        self.__results = results
        if verbose:
            print(f"특허 트래킹 완료.")
            print(f"첫 특허의 출원번호: {results[0][0]}")
            print(f"첫 특허의 결과 테이블 일부: {results[0][1][:3]}")

    async def track_patent(self, patent, verbose=False):
        """ Requests information of a patent and filter_out unneccesary information.
        
        Returns a 2-dimensional list.
        """
        records = await self.request_and_parse_kipris_API(application_number = patent, api_key = self.api_key, verbose=verbose)
        #print(f"records: {records}")
        result_table = await self.filter_records(records, verbose=verbose)
        # self.__result_dict[patent] = result_table
        return (patent, result_table)

    async def request_and_parse_kipris_API(self, application_number, api_key, verbose=False):
        """Request kipris REST API (asynchronously) and parse data using Beautifulsoup.

        soup.findall("relateddocsonfileInfo") will be returned.
        """
        
        url = 'http://plus.kipris.or.kr/openapi/rest/RelatedDocsonfilePatService/relatedDocsonfileInfo'
        query = f'?applicationNumber={application_number}&accessKey={api_key}'
        
        time1 = time.time()
        if verbose: print(f"request for patent:{application_number} started.")
        
        ## request by requests and loop.run_in_executor
        # import requests
        # loop_ = asyncio.get_event_loop()
        # response = await loop_.run_in_executor(None, requests.get, url+query)
        # text= response.text
        
        # request by aiohttp
        async with aiohttp.ClientSession() as session: 
            async with session.get(url+query) as response:
                text = await response.text()
        
        time2 = time.time()
        if verbose: print(f"request for patent:{application_number} finished. time:{time2-time1}")
        
        # parse
        soup = BeautifulSoup(text, "xml")
        records = soup.find_all('relateddocsonfileInfo')
        if records == []: 
            print("No records detected. Please check result message ")
            print(f"result_message: {soup.find('resultMsg').text}")

        return records

    async def filter_records(self, records, verbose=False):
        """ Filters out unnecessary records and fields.

        Returns a 2-dimensional list.
        """
        filtered_records = []
        time1 = time.time()
        for i, record in enumerate(records):
            ymd = record.documentDate.text
            record_date = date(int(ymd[:4]), int(ymd[4:6]), int(ymd[6:8]))
            if record_date < self.since or record_date > self.before:
                continue
            else:
                filtered_records.append([
                    i+1, #n-th record
                    record.documentTitle.text, #서류명
                    record.documentDate.text, #접수/발송 일자
                    record.status.text, #처리상태
                    record.step.text, #단계 (출원/등록 등.)
                    record.trialNumber.text, #심판 번호
                    record.registrationNumber.text #등록 번호
                ])
        time2 = time.time()
        if verbose: print(f"filtering records from a patent finished. time:{time2-time1}")
        return filtered_records
    
    ############### OUTPUT ###############
    def to_excel(self, verbose=False):
        """Saves result as an excel file(.xlsx)

        """
        print(self.results)
        if self.results == []:
            print("결과 값이 없습니다. 엑셀파일을 생성하지 않고 종료합니다.")
            #print("No results exists. Execute self.track_patents() to get results ")
            return
        
        # Create excel file
        if verbose: print("엑셀 파일 작성을 시작합니다.")
        result_wb = openpyxl.Workbook()
        result_ws = result_wb.active
        result_ws.title = 'result'

        # Apply sheet_style
        for letter in string.ascii_uppercase[:6]:
            result_ws.column_dimensions[letter] = sheet_style[f"col_{letter}_width"]

        current_row = 1

        # Write data
        for result in self.results:
            application_number = result[0]
            result_table = result[1]
            self._write_title(result_ws, current_row, title="출원번호: "+ application_number)
            current_row += 1
            self._write_info(result_ws, current_row, additional_info=self.additional_info_dict[application_number])
            current_row += 1
            self._write_fields(result_ws, current_row)
            current_row += 1
            self._write_records(result_ws, current_row, records=result_table)
            current_row += len(result_table)+2

            # print(f"출원번호 {application_number} 의 결과테이블 작성 완료.")
        
        #Save
        timestamp = time.strftime("%y%m%d_%H%M%S")
        output_name = output_directory + f"/output_{timestamp}.xlsx"
        result_wb.save(output_name)
        if verbose: print(f'엑셀 파일 {output_name}  저장을 완료했습니다.')

    def _write_title(self, result_ws, current_row, title):
        # default_title: application number.
        result_ws.merge_cells(f'A{current_row}:F{current_row}')
        result_ws[f'A{current_row}'].value = title
        result_ws[f'A{current_row}'].style = info_style

    def _write_info(self, result_ws, current_row, additional_info):

        #result_ws[f'A{current_row}'].value = info_0 
        for i,j in enumerate('BCDEF'):
            result_ws[f'{j}{current_row}'].value = additional_info[i] #info: from input.xlsx
        for row in result_ws[f"A{current_row}":f"F{current_row}"]:
            for cell in row:
                cell.style = info_style
    
    def _write_fields(self, result_ws, current_row):
        fields = ["번호", "서류명", "접수/발송일자", "처리단계", "단계", "심판/등록 번호"]
        for i,j in zip(fields, 'ABCDEF'):
            result_ws[f'{j}{current_row}'].value = i
        for row in result_ws[f"A{current_row}":f"F{current_row}"]:
            for cell in row:
                cell.style=field_style

    # records = 2D array (list) n*5
    def _write_records(self, result_ws, current_row, records):
        for row in records:
            number, document_title, document_date, status, step, trial_number, registration_number = row
            for i,j in zip(row[:5],'ABCDE'): #번호, 서류명, 접수/발송일자, 처리상태, 단계
                result_ws[f'{j}{current_row}'].value = i
            if trial_number !=' ':
                result_ws[f'F{current_row}'].value = trial_number #심판번호
            elif registration_number !=' ':
                result_ws[f'F{current_row}'].value = registration_number #등록번호           
       
            for row in result_ws[f"A{current_row}":f"F{current_row}"]:
                for cell in row:
                    cell.style=record_style
            current_row += 1

class DrugPatentTracker(PatentTracker):
    #override __write_info 
    def _write_info(self, result_ws, current_row, additional_info):
        item, authorization_holder, patent_holder, patent_class, patent_number = additional_info
        result_ws[f'B{current_row}'].value = item #품목
        result_ws[f'C{current_row}'].value = authorization_holder+"/"+patent_holder #허가권자/특허권자
        result_ws[f'D{current_row}'].value = patent_class #특허구분
        result_ws[f'F{current_row}'].value = patent_number #특허번호
        for row in result_ws[f"A{current_row}":f"F{current_row}"]:
            for cell in row:
                cell.style = info_style
    
    #as __write_info was overrode, to_excel also needs to be overidden
    def to_excel(self, verbose=False):
        """Saves result as an excel file(.xlsx)

        """
        
        if self.results == []:
            print("결과 값이 없습니다. 엑셀파일을 생성하지 않고 종료합니다.")
            #print("No results exists. Execute self.track_patents() to get results ")
            return
        
        # Create excel file
        if verbose: print("엑셀 파일 작성을 시작합니다.")
        result_wb = openpyxl.Workbook()
        result_ws = result_wb.active
        result_ws.title = 'result'

        # Apply sheet_style
        for letter in string.ascii_uppercase[:6]:
            result_ws.column_dimensions[letter].width = sheet_style[f"col_{letter}_width"]
        current_row = 1

        # Write data
        for result in self.results:
            application_number = result[0]
            result_table = result[1]
            super()._write_title(result_ws, current_row, title=application_number)
            current_row += 1
            self._write_info(result_ws, current_row, additional_info=self.additional_info_dict[application_number])
            current_row += 1
            super()._write_fields(result_ws, current_row)
            current_row += 1
            super()._write_records(result_ws, current_row, records=result_table)
            current_row += len(result_table)+2

            # print(f"출원번호 {application_number} 의 결과테이블 작성 완료.")
        
        #Save
        timestamp = time.strftime("%y%m%d_%H%M%S")
        output_name = output_directory + f"/output_{timestamp}.xlsx"
        result_wb.save(output_name)
        if verbose: print(f'엑셀 파일 {output_name}  저장을 완료했습니다.')

if __name__ == "__main__":
    # tracker = PatentTracker()
    wb = openpyxl.load_workbook(input_path) #Read xl file
    ws = wb['output_type']
    ouput_type = ws['C4'].value.strip().upper()
    if ouput_type == "DRUG":
        tracker = DrugPatentTracker()
    elif False: #ouput_type == "SEMICONDUCTOR"
        pass
    else: #ouput_type == "NORMAL"
        tracker = PatentTracker()
    time1 = time.time()
    loop = asyncio.get_event_loop()
    loop.run_until_complete(tracker.track_patents(verbose=True))
    loop.close
    time2 = time.time()
    print(f"total time taken: {time2-time1}")
    tracker.to_excel(verbose=True)